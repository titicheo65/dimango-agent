# -*- coding: utf-8 -*-
"""Cuadrante mensual SOLO MALL — Dimango. 29-jun a 26-jul 2026 (4 semanas).

Reglas aplicadas:
- 5 garzones, TODOS 42h exactas/semana (6x1, turnos de 7h). No se regalan ni exceden horas.
- Cobertura Mall: Lun-Jue 3, Vie-Sáb 4, Dom 3 (bajada de 4 a 3 para poder dar domingos libres).
- Los 4 "core" (Deyanira, Khrisbell, Paula, Sofía) reciben 2 DOMINGOS LIBRES c/u.
- María = flex (viene de Playa): trabaja todos los domingos en Mall (0 domingos libres este mes,
  rota el próximo) y absorbe el excedente -> se la envía a Playa (PLY). Sólo María va a Playa;
  sólo en días puntuales se manda un 2º core cuando ella está libre.
- Turnos rotativos AP/IN/CI.
"""

import collections
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ───────────────────────── Configuración base ─────────────────────────
INICIO = date(2026, 6, 29)
DIAS = 28
FECHAS = [INICIO + timedelta(d) for d in range(DIAS)]
WD = [f.weekday() for f in FECHAS]            # 0=Lun ... 6=Dom
DOM_IDX = [i for i in range(DIAS) if WD[i] == 6]
SEMANA = [i // 7 for i in range(DIAS)]
DIAS_ABBR = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]

CORE = ["Deyanira", "Khrisbell", "Paula", "Sofía"]
FLEX = "María"
MALL = CORE + [FLEX]
H = {"AP": 7, "IN": 7, "CI": 7, "PLY": 7, "LIB": 0}

def req_mall(wd):
    if wd <= 3:        # Lun-Jue
        return 3
    if wd in (4, 5):   # Vie, Sáb
        return 4
    return 3           # Dom (bajado de 4 a 3 para habilitar domingos libres)

# ───────────────────────── Plan de domingos libres (2 por core) ─────────────────────────
DOM_OFF = {
    DOM_IDX[0]: ["Deyanira", "Khrisbell"],   # 05/07
    DOM_IDX[1]: ["Paula", "Sofía"],          # 12/07
    DOM_IDX[2]: ["Deyanira", "Paula"],       # 19/07
    DOM_IDX[3]: ["Khrisbell", "Sofía"],      # 26/07
}

# ───────────────────────── Matriz de libres (1 libre/semana = 42h) ─────────────────────────
off = {w: [False] * DIAS for w in MALL}
for di, lst in DOM_OFF.items():
    for w in lst:
        off[w][di] = True

def asignar_libres():
    """Cada persona descansa exactamente 1 día por semana.
    Quien libra domingo ya tiene su descanso; el resto recibe 1 libre entre Lun-Jue
    (días de menor carga), en días distintos para no romper la cobertura."""
    weekday_slots = [0, 1, 2, 3]   # Lun-Jue
    for s in range(4):
        ya = [w for w in MALL if any(off[w][s * 7 + d] for d in range(7))]
        faltan = [w for w in MALL if w not in ya]
        for k, w in enumerate(faltan):
            wd_target = weekday_slots[(k + s) % len(weekday_slots)]
            off[w][s * 7 + wd_target] = True

asignar_libres()

# ───────────────────────── Turnos + envío a Playa (PLY) ─────────────────────────
grid = {w: [""] * DIAS for w in MALL}
core_ply = collections.Counter()
for idx in range(DIAS):
    d = req_mall(WD[idx])
    present = [w for w in MALL if not off[w][idx]]
    libres = [w for w in MALL if off[w][idx]]
    surplus = len(present) - d
    pres = list(present)
    ply = []
    # 1º a Playa: María (preferencia del cliente)
    if surplus > 0 and FLEX in pres:
        ply.append(FLEX); pres.remove(FLEX); surplus -= 1
    # 2º a Playa sólo si aún sobra (María libre ese día): rota el core menos usado
    while surplus > 0:
        cand = sorted([w for w in pres if w in CORE], key=lambda w: (core_ply[w], MALL.index(w)))
        if not cand:
            break
        sec = cand[0]; ply.append(sec); core_ply[sec] += 1; pres.remove(sec); surplus -= 1
    # turnos rotativos para los que quedan en Mall
    rot = idx % max(1, len(pres))
    orden = pres[rot:] + pres[:rot]
    pool = ["AP", "IN", "CI", "CI"][:d]
    for j, w in enumerate(orden):
        grid[w][idx] = pool[j] if j < len(pool) else "CI"
    for w in ply:
        grid[w][idx] = "PLY"
    for w in libres:
        grid[w][idx] = "LIB"

# ───────────────────────── Estilos ─────────────────────────
FUENTE = "Arial"
def F(**kw): kw.setdefault("name", FUENTE); return Font(**kw)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
C_HEAD = PatternFill("solid", fgColor="1F4E78")
C_DOM = PatternFill("solid", fgColor="FCE4D6")
C_LIB = PatternFill("solid", fgColor="D9D9D9")
C_TITLE = PatternFill("solid", fgColor="2E75B6")
C_ALERT = PatternFill("solid", fgColor="FFF2CC")
C_BAD = PatternFill("solid", fgColor="F8CBAD")
C_OK = PatternFill("solid", fgColor="C6E0B4")
C_SUB = PatternFill("solid", fgColor="DDEBF7")
TURNO_FILL = {
    "AP": PatternFill("solid", fgColor="FFF2CC"),
    "IN": PatternFill("solid", fgColor="DDEBF7"),
    "CI": PatternFill("solid", fgColor="E2EFDA"),
    "PLY": PatternFill("solid", fgColor="FFD9CC"),
    "LIB": C_LIB,
}

wb = Workbook()
def col(idx): return get_column_letter(idx)
def estilo_header(ws, row, ncols, fill=C_HEAD):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill; cell.font = F(bold=True, color="FFFFFF", size=10)
        cell.alignment = CENTER; cell.border = BORDER

# ───────────────────────── Cuadrante Mall ─────────────────────────
ws = wb.active; ws.title = "Mall - Cuadrante"
ncols = 1 + DIAS
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
t = ws.cell(row=1, column=1, value="DIMANGO MALL — Cuadrante 29/06 a 26/07 2026")
t.fill = C_TITLE; t.font = F(bold=True, color="FFFFFF", size=13); t.alignment = CENTER
ws.cell(row=2, column=1, value="Garzón"); ws.cell(row=3, column=1, value="Día")
for i in range(DIAS):
    c = i + 2
    fcell = ws.cell(row=2, column=c, value=FECHAS[i].strftime("%d/%m"))
    dcell = ws.cell(row=3, column=c, value=DIAS_ABBR[WD[i]])
    fcell.font = F(bold=True, color="FFFFFF", size=9); fcell.alignment = CENTER; fcell.fill = C_HEAD
    dcell.fill = C_DOM if WD[i] == 6 else C_HEAD
    dcell.font = F(bold=True, size=9, color=("9C0006" if WD[i] == 6 else "FFFFFF"))
    dcell.alignment = CENTER; fcell.border = BORDER; dcell.border = BORDER
for rr in (2, 3):
    cc = ws.cell(row=rr, column=1)
    cc.fill = C_HEAD; cc.font = F(bold=True, color="FFFFFF"); cc.alignment = CENTER; cc.border = BORDER
r = 4
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
sc = ws.cell(row=r, column=1, value="EQUIPO BASE (todos 42h/semana · 6x1)")
sc.fill = C_SUB; sc.font = F(bold=True, size=10); sc.alignment = LEFT
r += 1
for w in MALL:
    etiqueta = w + ("  (flex→Playa)" if w == FLEX else "")
    ws.cell(row=r, column=1, value=etiqueta).font = F(bold=True, size=10)
    ws.cell(row=r, column=1).alignment = LEFT; ws.cell(row=r, column=1).border = BORDER
    for i in range(DIAS):
        v = grid[w][i]
        cell = ws.cell(row=r, column=i + 2, value=v)
        cell.alignment = CENTER; cell.border = BORDER; cell.font = F(size=9)
        cell.fill = TURNO_FILL.get(v, PatternFill())
        if WD[i] == 6 and v == "LIB":
            cell.font = F(size=9, bold=True, color="375623")
    r += 1
ws.column_dimensions["A"].width = 18
for i in range(DIAS):
    ws.column_dimensions[col(i + 2)].width = 5.5
ws.freeze_panes = "B4"

# ───────────────────────── Turnos (leyenda) ─────────────────────────
ws_t = wb.create_sheet("Turnos")
ws_t.cell(row=1, column=1, value="LEYENDA DE TURNOS — MALL").font = F(bold=True, size=13, color="1F4E78")
hdr = ["Código", "Turno", "Horario", "Horas", "Notas"]
for c, h in enumerate(hdr, 1): ws_t.cell(row=3, column=c, value=h)
estilo_header(ws_t, 3, len(hdr))
turnos_data = [
    ["AP", "Apertura", "09:30 – 17:00", 7, "7h + 30 min colación"],
    ["IN", "Intermedio", "13:00 – 20:30", 7, "7h + 30 min colación"],
    ["CI", "Cierre", "15:00 – 22:30", 7, "7h + 30 min colación"],
    ["PLY", "Apoyo a Playa", "según turno Playa", 7, "Garzón sobrante de Mall enviado a Playa ese día (prefer. María)"],
    ["LIB", "Día libre", "—", 0, "Descanso (cada garzón 1 por semana)"],
]
for ri, row in enumerate(turnos_data, 4):
    for c, val in enumerate(row, 1):
        cell = ws_t.cell(row=ri, column=c, value=val)
        cell.border = BORDER; cell.font = F(size=10)
        cell.alignment = CENTER if c in (1, 4) else LEFT
        if c == 1: cell.fill = TURNO_FILL.get(val, PatternFill())
for c, w in enumerate([8, 18, 18, 7, 55], 1): ws_t.column_dimensions[col(c)].width = w

# ───────────────────────── Resumen Horas ─────────────────────────
ws_h = wb.create_sheet("Resumen Horas")
ws_h.cell(row=1, column=1, value="RESUMEN DE HORAS SEMANALES — MALL").font = F(bold=True, size=13, color="1F4E78")
hdr = ["Garzón", "Tipo", "Sem 1", "Sem 2", "Sem 3", "Sem 4", "Total", "Prom/sem", "Alerta"]
for c, h in enumerate(hdr, 1): ws_h.cell(row=3, column=c, value=h)
estilo_header(ws_h, 3, len(hdr))

def horas_semana(valores, wk):
    return sum(H.get(v, 0) for v in valores[wk * 7:wk * 7 + 7])

r = 4
for w in MALL:
    ws_h.cell(row=r, column=1, value=w).font = F(bold=True, size=10)
    ws_h.cell(row=r, column=2, value="Base 6x1 (42h)")
    semanas = [horas_semana(grid[w], wk) for wk in range(4)]
    for wk in range(4): ws_h.cell(row=r, column=3 + wk, value=semanas[wk])
    total = sum(semanas)
    ws_h.cell(row=r, column=7, value=total).font = F(bold=True)
    ws_h.cell(row=r, column=8, value=round(total / 4, 1))
    if max(semanas) > 42 or min(semanas) < 42:
        alerta = "⚠ ≠42h"; bad = True
    else:
        alerta = "OK 42h"; bad = False
    ac = ws_h.cell(row=r, column=9, value=alerta)
    ac.fill = C_BAD if bad else C_OK; ac.font = F(bold=True, size=10)
    for c in range(1, 10):
        ws_h.cell(row=r, column=c).border = BORDER
        ws_h.cell(row=r, column=c).alignment = CENTER if c != 1 else LEFT
    r += 1
for c, wdt in enumerate([14, 14, 7, 7, 7, 7, 8, 9, 9], 1): ws_h.column_dimensions[col(c)].width = wdt

# ───────────────────────── Domingos Libres ─────────────────────────
ws_d = wb.create_sheet("Domingos Libres")
ws_d.cell(row=1, column=1, value="CONTROL DE DOMINGOS LIBRES — MALL (meta: 2 por trabajador)").font = F(bold=True, size=13, color="1F4E78")
hdr = ["Garzón"] + [FECHAS[di].strftime("%d/%m") for di in DOM_IDX] + ["Dom. libres", "Cumple (≥2)"]
for c, h in enumerate(hdr, 1): ws_d.cell(row=3, column=c, value=h)
estilo_header(ws_d, 3, len(hdr))
col_cuenta = 2 + len(DOM_IDX); col_cumple = col_cuenta + 1
r = 4
for w in MALL:
    ws_d.cell(row=r, column=1, value=w).font = F(bold=True, size=10); ws_d.cell(row=r, column=1).alignment = LEFT
    cuenta = 0
    for j, di in enumerate(DOM_IDX):
        v = grid[w][di]; libre = (v == "LIB")
        if libre: cuenta += 1
        cell = ws_d.cell(row=r, column=2 + j, value=("LIBRE" if libre else v))
        if libre: cell.fill = C_OK; cell.font = F(bold=True, size=10, color="375623")
    ws_d.cell(row=r, column=col_cuenta, value=cuenta).font = F(bold=True)
    cm = ws_d.cell(row=r, column=col_cumple, value=("✓" if cuenta >= 2 else "⚠ <2"))
    if cuenta < 2: cm.fill = C_BAD; cm.font = F(bold=True)
    else: cm.fill = C_OK; cm.font = F(bold=True)
    for c in range(1, col_cumple + 1):
        ws_d.cell(row=r, column=c).border = BORDER
        if c != 1: ws_d.cell(row=r, column=c).alignment = CENTER
    r += 1
ws_d.cell(row=r + 1, column=1, value="María (viene de Playa) es el comodín del mes: trabaja los 4 domingos en Mall y absorbe el sobrante hacia Playa. Rota el próximo mes.").font = F(bold=True, color="9C0006", size=10)
ws_d.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=col_cumple)
ws_d.cell(row=r + 2, column=1, value="Para que TODOS (incl. María) tengan 2 domingos libres hay que bajar la cobertura dominical a 2 en 2 de los 4 domingos, o sumar un 6º garzón.").font = F(italic=True, size=10)
ws_d.merge_cells(start_row=r + 2, start_column=1, end_row=r + 2, end_column=col_cumple)
ws_d.column_dimensions["A"].width = 14
for j in range(len(DOM_IDX)): ws_d.column_dimensions[col(2 + j)].width = 10
ws_d.column_dimensions[col(col_cuenta)].width = 12; ws_d.column_dimensions[col(col_cumple)].width = 12

# ───────────────────────── Cobertura Diaria ─────────────────────────
ws_c = wb.create_sheet("Cobertura Diaria")
ws_c.cell(row=1, column=1, value="COBERTURA DIARIA — MALL").font = F(bold=True, size=13, color="1F4E78")
hdr = ["Fecha", "Día", "Req.", "En Mall", "✓", "A Playa (PLY)", "Quién va a Playa"]
for c, h in enumerate(hdr, 1): ws_c.cell(row=3, column=c, value=h)
estilo_header(ws_c, 3, len(hdr))
r = 4
for i in range(DIAS):
    rm = req_mall(WD[i])
    en_mall = sum(1 for w in MALL if grid[w][i] in ("AP", "IN", "CI"))
    quien_ply = [w for w in MALL if grid[w][i] == "PLY"]
    ws_c.cell(row=r, column=1, value=FECHAS[i].strftime("%d/%m"))
    ws_c.cell(row=r, column=2, value=DIAS_ABBR[WD[i]])
    ws_c.cell(row=r, column=3, value=rm)
    ws_c.cell(row=r, column=4, value=en_mall)
    mc = ws_c.cell(row=r, column=5, value=("OK" if en_mall >= rm else "FALTA"))
    ws_c.cell(row=r, column=6, value=len(quien_ply))
    ws_c.cell(row=r, column=7, value=", ".join(quien_ply) if quien_ply else "—")
    mc.fill = C_OK if en_mall >= rm else C_BAD
    for c in range(1, 8):
        cell = ws_c.cell(row=r, column=c); cell.border = BORDER
        cell.alignment = CENTER if c != 7 else LEFT; cell.font = F(size=10)
        if WD[i] == 6 and c <= 2: cell.fill = C_DOM
    r += 1
for c, wdt in enumerate([9, 7, 7, 9, 7, 14, 26], 1): ws_c.column_dimensions[col(c)].width = wdt
ws_c.freeze_panes = "A4"

# ───────────────────────── Alertas ─────────────────────────
ws_a = wb.create_sheet("Alertas")
ws_a.cell(row=1, column=1, value="ALERTAS Y DECISIONES — MALL").font = F(bold=True, size=13, color="9C0006")
alertas = [
    ("OK", "Horas (LEY)", "Los 5 garzones trabajan 6x1 = 6 días × 7h = 42h EXACTAS cada semana. Ver 'Resumen Horas': todos en 42h, sin regalar ni exceder."),
    ("OK", "Domingos libres (core)", "Deyanira, Khrisbell, Paula y Sofía reciben 2 domingos libres c/u (cumplido)."),
    ("DECISIÓN", "Cobertura dominical bajó de 4 a 3", "Con 5 garzones y 42h obligatorias, dar domingos libres EXIGE soltar gente el domingo. Con cobertura 4 sólo cabía 1 libre por domingo (4 cupos para 5 personas). Bajando a 3 hay 8 cupos → 2 c/u para los 4 core. CONFIRMA si aceptas cobertura 3 los domingos."),
    ("PARCIAL", "María con 0 domingos libres", "Para 8 cupos y 5 personas, alguien queda con 0. Elegí a María porque viene de Playa y es el comodín que se envía allá. Rota el próximo mes. Si quieres que TODOS (incl. María) tengan 2, hay que bajar la cobertura dominical a 2 en 2 domingos, o sumar un 6º garzón."),
    ("OK", "Turnos rotativos", "AP/IN/CI y el día libre rotan por persona y semana. Nadie queda fijo en apertura ni cierre."),
    ("ATENCIÓN", "Sobrante a Playa = María", "Mall sobra-staff: necesita ~23 turnos/semana y 5 garzones aportan 30 → 7 turnos/semana sobran y van a Playa (PLY). Se concentran en María (la mayoría de su semana queda en Playa). Sólo días puntuales en que ella libra se manda un 2º core. Ver 'Cobertura Diaria' col. 'Quién va a Playa'."),
    ("NOTA", "Pendiente Playa", "Estos turnos PLY hay que encajarlos en la planilla de Playa (siguiente paso, cuando cuadremos Mall)."),
]
hdr = ["Nivel", "Tema", "Detalle"]
for c, h in enumerate(hdr, 1): ws_a.cell(row=3, column=c, value=h)
estilo_header(ws_a, 3, len(hdr))
nivel_fill = {"CRÍTICO": C_BAD, "ATENCIÓN": C_ALERT, "DECISIÓN": C_ALERT, "PARCIAL": C_ALERT, "OK": C_OK, "NOTA": C_SUB}
r = 4
for niv, tema, det in alertas:
    ws_a.cell(row=r, column=1, value=niv).fill = nivel_fill.get(niv, PatternFill())
    ws_a.cell(row=r, column=1).font = F(bold=True, size=10)
    ws_a.cell(row=r, column=2, value=tema).font = F(bold=True, size=10)
    ws_a.cell(row=r, column=3, value=det).font = F(size=10)
    for c in range(1, 4):
        ws_a.cell(row=r, column=c).border = BORDER
        ws_a.cell(row=r, column=c).alignment = LEFT if c == 3 else CENTER
    r += 1
ws_a.column_dimensions["A"].width = 11; ws_a.column_dimensions["B"].width = 26; ws_a.column_dimensions["C"].width = 100

# ───────────────────────── Resumen (portada) ─────────────────────────
ws_i = wb.create_sheet("Resumen", 0)
ws_i.sheet_view.showGridLines = False
ws_i.cell(row=1, column=1, value="CUADRANTE MENSUAL — DIMANGO MALL").font = F(bold=True, size=16, color="1F4E78")
ws_i.cell(row=2, column=1, value="Periodo: lunes 29 de junio a domingo 26 de julio de 2026 (4 semanas)").font = F(size=11)
lineas = [
    "", "HOJAS DE ESTE ARCHIVO:",
    "• Mall - Cuadrante   → grilla diaria de los 5 garzones.",
    "• Turnos             → leyenda de códigos, horarios y horas.",
    "• Resumen Horas      → horas por semana (todos en 42h exactas).",
    "• Domingos Libres    → control de domingos libres por trabajador.",
    "• Cobertura Diaria   → requerido vs en Mall y quién se envía a Playa.",
    "• Alertas            → decisiones y supuestos. LEER PRIMERO.",
    "", "REGLAS APLICADAS:",
    "• Todos 42h exactas/semana (6x1, turnos de 7h). Es ley: no se regalan ni exceden horas.",
    "• Cobertura: Lun-Jue 3 · Vie-Sáb 4 · Dom 3 (bajada de 4 para dar domingos libres).",
    "• 4 core con 2 domingos libres c/u. María = comodín (0 este mes, rota el próximo).",
    "• Turnos rotativos. Sobrante de Mall → Playa (PLY), concentrado en María.",
    "", "EQUIPO MALL: Deyanira, Khrisbell, Paula, Sofía, María.",
    "", "⚠ Ver 'Alertas': hay que CONFIRMAR la cobertura dominical de 3. Playa es el siguiente paso.",
]
r = 4
for ln in lineas:
    cell = ws_i.cell(row=r, column=1, value=ln)
    if ln.endswith(":"): cell.font = F(bold=True, size=11, color="1F4E78")
    elif ln.startswith("⚠"): cell.font = F(bold=True, size=11, color="9C0006")
    else: cell.font = F(size=10)
    r += 1
ws_i.column_dimensions["A"].width = 105

orden = ["Resumen", "Mall - Cuadrante", "Turnos", "Resumen Horas",
         "Domingos Libres", "Cobertura Diaria", "Alertas"]
wb._sheets.sort(key=lambda s: orden.index(s.title) if s.title in orden else 99)

OUT = "/Users/ricardovinet/whatsapp-agentkit/Cuadrante_Mall_Dimango_Jun-Jul_2026.xlsx"
wb.save(OUT)
print("Guardado:", OUT)

# ───────────────────────── Validación ─────────────────────────
print("\n--- Validación MALL ---")
ok = True
for w in MALL:
    for wk in range(4):
        h = horas_semana(grid[w], wk)
        if h != 42:
            print(f"{w} sem{wk+1}: {h}h <-- ! (debe ser 42)"); ok = False
for w in MALL:
    dl = sum(1 for di in DOM_IDX if grid[w][di] == "LIB")
    print(f"{w}: domingos libres = {dl}")
for i in range(DIAS):
    en_mall = sum(1 for w in MALL if grid[w][i] in ("AP", "IN", "CI"))
    if en_mall < req_mall(WD[i]):
        print(f"{FECHAS[i].strftime('%d/%m')}: en Mall={en_mall} < req {req_mall(WD[i])} <-- !"); ok = False
ply_maria = sum(1 for i in range(DIAS) if grid[FLEX][i] == "PLY")
ply_core = sum(1 for w in CORE for i in range(DIAS) if grid[w][i] == "PLY")
print(f"PLY María={ply_maria}  PLY core={ply_core}  (total/sem≈{(ply_maria+ply_core)/4:.1f})")
print("Resultado:", "OK" if ok else "REVISAR")
