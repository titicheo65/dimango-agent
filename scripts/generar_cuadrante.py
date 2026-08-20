# -*- coding: utf-8 -*-
"""Generador de cuadrante mensual de garzones — Dimango Mall y Playa.
Periodo: lunes 29-jun-2026 a domingo 26-jul-2026 (4 semanas exactas).

Roster:
  Mall (5 base):  Deyanira, Khrisbell, Paula, Sofía, María
  Playa (6 base): Jesús, Rubén, Marco, Nicolás, Fabián, Samuel
  Playa PT (2):   Jhony, Juan
"""

from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ───────────────────────── Configuración base ─────────────────────────
INICIO = date(2026, 6, 29)
DIAS = 28
FECHAS = [INICIO + timedelta(d) for d in range(DIAS)]
WD = [f.weekday() for f in FECHAS]            # 0=Lun ... 6=Dom
DOM_IDX = [i for i in range(DIAS) if WD[i] == 6]
SEMANA = [i // 7 for i in range(DIAS)]
DIAS_ABBR = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]

MALL = ["Deyanira", "Khrisbell", "Paula", "Sofía", "María"]
PLAYA_FT = ["Jesús", "Rubén", "Marco", "Nicolás", "Fabián", "Samuel"]
PLAYA_PT = ["Jhony", "Juan"]

# Horas por código de turno
MALL_H = {"AP": 7, "IN": 7, "CI": 7, "PLY": 7, "LIB": 0}
PLAYA_H = {"AP": 7, "IN": 7, "CI": 7, "LIB": 0}
PT_H = {"PA": 9, "PB": 8, "LIB": 0}

# ───────────────────────── Cobertura requerida ─────────────────────────
def req_mall(wd):
    return 3 if wd <= 3 else 4

def req_playa_ft(wd):
    # Domingo: 3 FT (los otros 2 base los cubren los part-time). Resto: 5 FT.
    return 3 if wd == 6 else 5

# ───────────────────────── Plan de domingos libres ─────────────────────────
# Playa: 6 FT, 3 libres por domingo -> 12 cupos = 2 por trabajador.
DOM_OFF_PLAYA = {
    DOM_IDX[0]: ["Jesús", "Rubén", "Marco"],        # 05/07
    DOM_IDX[1]: ["Nicolás", "Fabián", "Samuel"],    # 12/07
    DOM_IDX[2]: ["Jesús", "Nicolás", "Fabián"],     # 19/07
    DOM_IDX[3]: ["Rubén", "Marco", "Samuel"],       # 26/07
}
# ───────────────────────── Scheduler genérico de libres ─────────────────────────
def schedule_local(workers, req_func, dom_plan):
    """Devuelve dict worker -> [bool libre por día] cumpliendo cobertura exacta."""
    off = {w: [False] * DIAS for w in workers}
    for di, lst in dom_plan.items():
        for w in lst:
            off[w][di] = True
    for s in range(4):
        for d in range(7):
            idx = s * 7 + d
            need_off = len(workers) - req_func(WD[idx])
            ya = sum(1 for w in workers if off[w][idx])
            faltan = need_off - ya
            if faltan <= 0:
                continue
            cand = [w for w in workers if not off[w][idx]]
            week_off = {w: sum(off[w][s*7:s*7+7]) for w in cand}
            total_off = {w: sum(off[w]) for w in cand}
            cand.sort(key=lambda w: (week_off[w], total_off[w], workers.index(w)))
            for w in cand[:faltan]:
                off[w][idx] = True
    return off

playa_off = schedule_local(PLAYA_FT, req_playa_ft, DOM_OFF_PLAYA)

# ───────────────────────── Mall: 6x1 = 42h exactas, rotativo, sobrante a Playa ─────────────────────────
def build_mall():
    """Mall con 5 garzones, todos 6x1 = 6 días × 7h = 42h exactas (ley).
    Cada día se cubre la demanda (3 L-J / 4 V-D) y el sobrante se marca PLY
    (garzón enviado a apoyar Playa). El único día libre rota por persona y
    semana, incluyendo el domingo, para repartir los domingos libres."""
    workers = MALL
    n = len(workers)            # 5
    dayslots = [0, 1, 2, 3, 6]  # Lun, Mar, Mié, Jue, Dom -> días candidatos a libre
    off = {w: [False] * DIAS for w in workers}
    for s in range(4):
        for i, w in enumerate(workers):
            slot_wd = dayslots[(i + s) % n]     # rota el día libre por semana
            for d in range(7):
                idx = s * 7 + d
                if WD[idx] == slot_wd:
                    off[w][idx] = True
                    break
    grid = {w: [""] * DIAS for w in workers}
    for idx in range(DIAS):
        demand = req_mall(WD[idx])
        trabajan = [w for w in workers if not off[w][idx]]
        libres = [w for w in workers if off[w][idx]]
        rot = idx % max(1, len(trabajan))       # turnos rotativos
        orden = trabajan[rot:] + trabajan[:rot]
        turnos = ["AP", "IN", "CI", "CI"][:demand]
        for j, w in enumerate(orden):
            grid[w][idx] = turnos[j] if j < demand else "PLY"   # sobrante -> Playa
        for w in libres:
            grid[w][idx] = "LIB"
    return off, grid

mall_off, mall_grid = build_mall()

# ───────────────────────── Turnos FT ─────────────────────────
def asignar_turnos(workers, off, pool_func):
    grid = {w: [""] * DIAS for w in workers}
    for i in range(DIAS):
        trabajan = [w for w in workers if not off[w][i]]
        libres = [w for w in workers if off[w][i]]
        pool = pool_func(WD[i], len(trabajan))
        rot = i % max(1, len(trabajan))
        orden = trabajan[rot:] + trabajan[:rot]
        for j, w in enumerate(orden):
            grid[w][i] = pool[j] if j < len(pool) else "CI"
        for w in libres:
            grid[w][i] = "LIB"
    return grid

def pool_playa(wd, n):
    return ["AP", "IN"] + ["CI"] * max(0, n - 2)

playa_grid = asignar_turnos(PLAYA_FT, playa_off, pool_playa)

# ───────────────────────── Part-time Playa ─────────────────────────
pt_grid = {w: ["LIB"] * DIAS for w in PLAYA_PT}
for i in range(DIAS):
    if WD[i] == 4:      # Viernes: 1 apoyo
        pt_grid[PLAYA_PT[SEMANA[i] % 2]][i] = "PA"
    elif WD[i] == 5:    # Sábado: 1 apoyo
        pt_grid[PLAYA_PT[(SEMANA[i] + 1) % 2]][i] = "PA"
for di in DOM_IDX:
    n_off = len(DOM_OFF_PLAYA.get(di, []))
    ft_trab = len(PLAYA_FT) - n_off
    falta_base = max(0, 5 - ft_trab)
    asignados = 0
    for pt in PLAYA_PT:
        if asignados < falta_base:
            pt_grid[pt][di] = "PB"
            asignados += 1
    for pt in PLAYA_PT:
        if pt_grid[pt][di] == "LIB" and falta_base < 2:
            pt_grid[pt][di] = "PA"
            break

# ───────────────────────── Estilos ─────────────────────────
FUENTE = "Arial"
def F(**kw): kw.setdefault("name", FUENTE); return Font(**kw)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
C_HEAD = PatternFill("solid", fgColor="1F4E78")
C_DOM = PatternFill("solid", fgColor="FCE4D6")
C_LIB = PatternFill("solid", fgColor="D9D9D9")
C_PT = PatternFill("solid", fgColor="E2EFDA")
C_TITLE = PatternFill("solid", fgColor="2E75B6")
C_ALERT = PatternFill("solid", fgColor="FFF2CC")
C_BAD = PatternFill("solid", fgColor="F8CBAD")
C_OK = PatternFill("solid", fgColor="C6E0B4")
C_SUB = PatternFill("solid", fgColor="DDEBF7")
TURNO_FILL = {
    "AP": PatternFill("solid", fgColor="FFF2CC"),
    "IN": PatternFill("solid", fgColor="DDEBF7"),
    "CI": PatternFill("solid", fgColor="E2EFDA"),
    "PLY": PatternFill("solid", fgColor="FFD9CC"),
    "PA": PatternFill("solid", fgColor="D9E1F2"),
    "PB": PatternFill("solid", fgColor="D9E1F2"),
    "LIB": C_LIB,
}

wb = Workbook()
def col(idx): return get_column_letter(idx)
def estilo_header(ws, row, ncols, fill=C_HEAD):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill; cell.font = F(bold=True, color="FFFFFF", size=10)
        cell.alignment = CENTER; cell.border = BORDER

# ───────────────────────── Cuadrante ─────────────────────────
def cuadrante_sheet(ws, titulo, base_workers, base_grid,
                    pt_workers=None, pt_grid=None):
    ncols = 1 + DIAS
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(row=1, column=1, value=titulo)
    t.fill = C_TITLE; t.font = F(bold=True, color="FFFFFF", size=13); t.alignment = CENTER
    ws.cell(row=2, column=1, value="Garzón"); ws.cell(row=3, column=1, value="Día")
    for i in range(DIAS):
        c = i + 2
        fcell = ws.cell(row=2, column=c, value=FECHAS[i].strftime("%d/%m"))
        dcell = ws.cell(row=3, column=c, value=DIAS_ABBR[WD[i]])
        fcell.font = F(bold=True, color="FFFFFF", size=9); fcell.alignment = CENTER; fcell.fill = C_HEAD
        dcell.fill = C_DOM if WD[i] == 6 else C_HEAD
        dcell.font = F(bold=True, size=9, color=("9C0006" if WD[i] == 6 else "FFFFFF"))
        dcell.alignment = CENTER; fcell.border = BORDER; dcell.border = BORDER
    for rr in (2, 3):
        cc = ws.cell(row=rr, column=1)
        cc.fill = C_HEAD; cc.font = F(bold=True, color="FFFFFF"); cc.alignment = CENTER; cc.border = BORDER
    r = 4
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    sc = ws.cell(row=r, column=1, value="EQUIPO BASE (full-time)")
    sc.fill = C_SUB; sc.font = F(bold=True, size=10); sc.alignment = LEFT
    r += 1
    base_start = r
    for w in base_workers:
        ws.cell(row=r, column=1, value=w).font = F(bold=True, size=10)
        ws.cell(row=r, column=1).alignment = LEFT; ws.cell(row=r, column=1).border = BORDER
        for i in range(DIAS):
            v = base_grid[w][i]
            cell = ws.cell(row=r, column=i+2, value=v)
            cell.alignment = CENTER; cell.border = BORDER; cell.font = F(size=9)
            cell.fill = TURNO_FILL.get(v, PatternFill())
            if WD[i] == 6 and v == "LIB":
                cell.font = F(size=9, bold=True, color="375623")
        r += 1
    base_end = r - 1
    pt_start = pt_end = None
    if pt_workers:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        sc = ws.cell(row=r, column=1, value="APOYO PART-TIME")
        sc.fill = C_PT; sc.font = F(bold=True, size=10); sc.alignment = LEFT
        r += 1
        pt_start = r
        for w in pt_workers:
            ws.cell(row=r, column=1, value=w).font = F(bold=True, size=10)
            ws.cell(row=r, column=1).alignment = LEFT; ws.cell(row=r, column=1).border = BORDER
            for i in range(DIAS):
                v = pt_grid[w][i]
                cell = ws.cell(row=r, column=i+2, value=v)
                cell.alignment = CENTER; cell.border = BORDER; cell.font = F(size=9)
                cell.fill = TURNO_FILL.get(v, PatternFill())
            r += 1
        pt_end = r - 1
    ws.column_dimensions["A"].width = 14
    for i in range(DIAS):
        ws.column_dimensions[col(i+2)].width = 5.5
    ws.freeze_panes = "B4"
    return base_start, base_end, pt_start, pt_end

ws_mall = wb.active; ws_mall.title = "Mall - Cuadrante"
m_bs, m_be, _, _ = cuadrante_sheet(ws_mall, "DIMANGO MALL — Cuadrante 29/06 a 26/07 2026", MALL, mall_grid)
ws_playa = wb.create_sheet("Playa - Cuadrante")
p_bs, p_be, p_ps, p_pe = cuadrante_sheet(ws_playa, "DIMANGO PLAYA — Cuadrante 29/06 a 26/07 2026",
                                         PLAYA_FT, playa_grid, PLAYA_PT, pt_grid)

# ───────────────────────── Turnos (leyenda) ─────────────────────────
ws_t = wb.create_sheet("Turnos")
ws_t.cell(row=1, column=1, value="LEYENDA DE TURNOS").font = F(bold=True, size=13, color="1F4E78")
hdr = ["Local", "Código", "Turno", "Horario", "Horas", "Notas"]
for c, h in enumerate(hdr, 1): ws_t.cell(row=3, column=c, value=h)
estilo_header(ws_t, 3, len(hdr))
turnos_data = [
    ["Mall", "AP", "Apertura", "09:30 – 17:00", 7, "7h + 30 min colación"],
    ["Mall", "IN", "Intermedio", "13:00 – 20:30", 7, "7h + 30 min colación"],
    ["Mall", "CI", "Cierre", "15:00 – 22:30", 7, "7h + 30 min colación"],
    ["Mall", "PLY", "Apoyo a Playa", "según turno Playa", 7, "Garzón sobrante de Mall enviado a Playa ese día"],
    ["Playa", "AP", "Apertura", "11:30 – 19:00", 7, ""],
    ["Playa", "IN", "Intermedio", "14:00 – 21:30", 7, ""],
    ["Playa", "CI", "Cierre (L-J)", "17:00 – 00:30", 7, "Vie/Sáb 18:00–01:30 · Dom 16:30–00:00"],
    ["Playa", "PA", "Apoyo part-time", "≈16:00 – 01:30", 9, "Refuerzo Vie/Sáb/Dom (8–10h)"],
    ["Playa", "PB", "PT base (backfill)", "≈16:30 – 00:30", 8, "PT cubre base cuando FT libra domingo"],
    ["Ambos", "LIB", "Día libre", "—", 0, "Descanso"],
]
for ri, row in enumerate(turnos_data, 4):
    for c, val in enumerate(row, 1):
        cell = ws_t.cell(row=ri, column=c, value=val)
        cell.border = BORDER; cell.font = F(size=10)
        cell.alignment = CENTER if c in (2, 5) else LEFT
        if c == 2: cell.fill = TURNO_FILL.get(val, PatternFill())
for c, w in enumerate([8, 8, 18, 16, 7, 42], 1): ws_t.column_dimensions[col(c)].width = w

# ───────────────────────── Resumen Horas ─────────────────────────
ws_h = wb.create_sheet("Resumen Horas")
ws_h.cell(row=1, column=1, value="RESUMEN DE HORAS SEMANALES").font = F(bold=True, size=13, color="1F4E78")
hdr = ["Garzón", "Local", "Tipo", "Sem 1", "Sem 2", "Sem 3", "Sem 4", "Total", "Prom/sem", "Alerta"]
for c, h in enumerate(hdr, 1): ws_h.cell(row=3, column=c, value=h)
estilo_header(ws_h, 3, len(hdr))

def horas_semana(valores, wk, hmap):
    return sum(hmap.get(v, 0) for v in valores[wk*7:wk*7+7])

r = 4
def fila_horas(ws, r, nombre, local, tipo, valores, hmap):
    ws.cell(row=r, column=1, value=nombre).font = F(bold=True, size=10)
    ws.cell(row=r, column=2, value=local); ws.cell(row=r, column=3, value=tipo)
    semanas = [horas_semana(valores, wk, hmap) for wk in range(4)]
    for wk in range(4): ws.cell(row=r, column=4+wk, value=semanas[wk])
    total = sum(semanas)
    ws.cell(row=r, column=8, value=total).font = F(bold=True)
    ws.cell(row=r, column=9, value=round(total/4, 1))
    alerta = "⚠ >42h" if max(semanas) > 42 else "OK"
    ac = ws.cell(row=r, column=10, value=alerta)
    if alerta != "OK": ac.fill = C_BAD; ac.font = F(bold=True, size=10)
    for c in range(1, 11):
        ws.cell(row=r, column=c).border = BORDER
        if ws.cell(row=r, column=c).alignment.horizontal is None:
            ws.cell(row=r, column=c).alignment = CENTER
    return r+1

for w in MALL:
    r = fila_horas(ws_h, r, w, "Mall", "Base 6x1 (42h)", mall_grid[w], MALL_H)
for w in PLAYA_FT:
    r = fila_horas(ws_h, r, w, "Playa", "Base 6x1/5x2", playa_grid[w], PLAYA_H)
for w in PLAYA_PT:
    r = fila_horas(ws_h, r, w, "Playa", "Part-time", pt_grid[w], PT_H)
for c, wdt in enumerate([14, 8, 12, 7, 7, 7, 7, 8, 9, 9], 1): ws_h.column_dimensions[col(c)].width = wdt

# ───────────────────────── Domingos Libres ─────────────────────────
ws_d = wb.create_sheet("Domingos Libres")
ws_d.cell(row=1, column=1, value="CONTROL DE DOMINGOS LIBRES (meta: 2 por trabajador)").font = F(bold=True, size=13, color="1F4E78")
hdr = ["Garzón", "Local"] + [FECHAS[di].strftime("%d/%m") for di in DOM_IDX] + ["Domingos libres", "Cumple (≥2)"]
for c, h in enumerate(hdr, 1): ws_d.cell(row=3, column=c, value=h)
estilo_header(ws_d, 3, len(hdr))
col_cuenta = 3 + len(DOM_IDX); col_cumple = col_cuenta + 1

def fila_domingos(ws, r, nombre, local, grid):
    ws.cell(row=r, column=1, value=nombre).font = F(bold=True, size=10); ws.cell(row=r, column=1).alignment = LEFT
    ws.cell(row=r, column=2, value=local)
    cuenta = 0
    for j, di in enumerate(DOM_IDX):
        v = grid[nombre][di]; libre = (v == "LIB")
        if libre: cuenta += 1
        cell = ws.cell(row=r, column=3+j, value=("LIBRE" if libre else v))
        if libre: cell.fill = C_OK; cell.font = F(bold=True, size=10, color="375623")
    ws.cell(row=r, column=col_cuenta, value=cuenta).font = F(bold=True)
    cm = ws.cell(row=r, column=col_cumple, value=("✓" if cuenta >= 2 else "⚠ <2"))
    if cuenta < 2: cm.fill = C_BAD; cm.font = F(bold=True)
    for c in range(1, col_cumple+1):
        ws.cell(row=r, column=c).border = BORDER
        if c != 1: ws.cell(row=r, column=c).alignment = CENTER
    return r+1

r = 4
for w in PLAYA_FT:
    r = fila_domingos(ws_d, r, w, "Playa", playa_grid)
for w in MALL:
    r = fila_domingos(ws_d, r, w, "Mall", mall_grid)
ws_d.cell(row=r+1, column=1, value="MALL: con 5 garzones y cobertura de 4 los Vie-Dom solo se alcanza ~1 domingo libre por persona (rotación).").font = F(bold=True, color="9C0006", size=10)
ws_d.merge_cells(start_row=r+1, start_column=1, end_row=r+1, end_column=col_cumple)
ws_d.cell(row=r+2, column=1, value="Para 2 domingos libres c/u en Mall: bajar cobertura de finde a 3 algunos domingos o sumar un 6º garzón.").font = F(italic=True, size=10)
ws_d.merge_cells(start_row=r+2, start_column=1, end_row=r+2, end_column=col_cumple)
ws_d.column_dimensions["A"].width = 14; ws_d.column_dimensions["B"].width = 8
for j in range(len(DOM_IDX)): ws_d.column_dimensions[col(3+j)].width = 10
ws_d.column_dimensions[col(col_cuenta)].width = 15; ws_d.column_dimensions[col(col_cumple)].width = 12

# ───────────────────────── Cobertura Diaria ─────────────────────────
ws_c = wb.create_sheet("Cobertura Diaria")
ws_c.cell(row=1, column=1, value="COBERTURA DIARIA POR LOCAL").font = F(bold=True, size=13, color="1F4E78")
hdr = ["Fecha", "Día", "Mall req.", "Mall asig.", "Mall ✓", "Playa req. (base)", "Playa asig. (base)", "Playa ✓", "Playa apoyo PT"]
for c, h in enumerate(hdr, 1): ws_c.cell(row=3, column=c, value=h)
estilo_header(ws_c, 3, len(hdr))
r = 4
for i in range(DIAS):
    rm = req_mall(WD[i]); rp = 5
    mall_asig = sum(1 for w in MALL if mall_grid[w][i] in ("AP", "IN", "CI"))
    playa_base = sum(1 for w in PLAYA_FT if playa_grid[w][i] != "LIB") + sum(1 for w in PLAYA_PT if pt_grid[w][i] == "PB")
    playa_pa = sum(1 for w in PLAYA_PT if pt_grid[w][i] == "PA")
    ws_c.cell(row=r, column=1, value=FECHAS[i].strftime("%d/%m"))
    ws_c.cell(row=r, column=2, value=DIAS_ABBR[WD[i]])
    ws_c.cell(row=r, column=3, value=rm); ws_c.cell(row=r, column=4, value=mall_asig)
    mc = ws_c.cell(row=r, column=5, value=("OK" if mall_asig >= rm else "FALTA"))
    ws_c.cell(row=r, column=6, value=rp); ws_c.cell(row=r, column=7, value=playa_base)
    pc = ws_c.cell(row=r, column=8, value=("OK" if playa_base >= rp else "FALTA"))
    ws_c.cell(row=r, column=9, value=playa_pa)
    mc.fill = C_OK if mall_asig >= rm else C_BAD
    pc.fill = C_OK if playa_base >= rp else C_BAD
    for c in range(1, 10):
        cell = ws_c.cell(row=r, column=c); cell.border = BORDER; cell.alignment = CENTER; cell.font = F(size=10)
        if WD[i] == 6 and c <= 2: cell.fill = C_DOM
    r += 1
for c, wdt in enumerate([9, 7, 10, 11, 9, 16, 17, 9, 14], 1): ws_c.column_dimensions[col(c)].width = wdt
ws_c.freeze_panes = "A4"

# ───────────────────────── Alertas ─────────────────────────
ws_a = wb.create_sheet("Alertas")
ws_a.cell(row=1, column=1, value="ALERTAS Y SUPUESTOS DEL CUADRANTE").font = F(bold=True, size=13, color="9C0006")
alertas = [
    ("OK", "Mall — horas (LEY)", "Los 5 garzones trabajan 6x1 = 6 días × 7h = 42h EXACTAS cada semana. No se regalan ni se exceden horas. Ver hoja 'Resumen Horas': todos en 42h."),
    ("NOTA", "Mall — 42h exactas requieren 6x1", "Con turnos de 7h, 42h sólo se logra trabajando 6 días (6×7). Una semana de 5 días daría 35h (7h) o 40h (8h) — se regalarían horas; y meter un turno de 8h en una semana de 6 días daría 43h — se excede. Por eso todas las semanas son 6x1 de 7h."),
    ("OK", "Mall — turnos rotativos", "El turno (AP/IN/CI) y el día libre rotan por persona y por semana. Nadie queda fijo en apertura ni en cierre."),
    ("ATENCIÓN", "Mall — sobrante a Playa (PLY)", "Como la demanda de Mall (3 L-J / 4 V-D) es menor que los garzones disponibles, cada día sobra 1 garzón que se marca PLY y se envía a apoyar Playa: 6 turnos/semana (1 Lun-Sáb, 0 Dom). Falta redistribuir estos 6 turnos dentro de la planilla de Playa (pendiente: rebalance de Playa)."),
    ("PARCIAL", "Mall — domingos libres", "Con 5 garzones y cobertura de 4 los domingos, sólo 1 puede librar por domingo → ~1 domingo libre por persona en el periodo. Deyanira queda con 0 este mes (rota el próximo). Para 2 c/u: bajar cobertura dominical a 3 o sumar un 6º garzón."),
    ("OK", "Playa — domingos libres", "Los 6 garzones base reciben exactamente 2 domingos libres, con los 2 part-time cubriendo la base todos los domingos."),
    ("ATENCIÓN", "Playa — horas", "Con 6 base (antes 7) el equipo sube a 6x1/5x2 y algunas semanas llega a 42h (tope). Menos holgura para reemplazos. Considerar mantener 7 en Playa si hay reemplazos frecuentes."),
    ("NOTA", "Playa — apoyo dominical", "Como ambos part-time cubren base los domingos, esos días no queda un apoyo de refuerzo nocturno separado."),
    ("NOTA", "Part-time", "Jhony y Juan se asignan a Vie/Sáb/Dom (apoyo y backfill). Turnos 8–10h. Pueden trabajar pasada medianoche (cierres 00:30/01:30)."),
    ("REVISAR", "Cobertura", "Ver hoja 'Cobertura Diaria': cada día debe marcar OK en ambos locales."),
]
hdr = ["Nivel", "Tema", "Detalle"]
for c, h in enumerate(hdr, 1): ws_a.cell(row=3, column=c, value=h)
estilo_header(ws_a, 3, len(hdr))
nivel_fill = {"CRÍTICO": C_BAD, "ATENCIÓN": C_ALERT, "PARCIAL": C_ALERT, "AJUSTE": C_ALERT, "OK": C_OK, "NOTA": C_SUB, "REVISAR": C_ALERT}
r = 4
for niv, tema, det in alertas:
    ws_a.cell(row=r, column=1, value=niv).fill = nivel_fill.get(niv, PatternFill())
    ws_a.cell(row=r, column=1).font = F(bold=True, size=10)
    ws_a.cell(row=r, column=2, value=tema).font = F(bold=True, size=10)
    ws_a.cell(row=r, column=3, value=det).font = F(size=10)
    for c in range(1, 4):
        ws_a.cell(row=r, column=c).border = BORDER
        ws_a.cell(row=r, column=c).alignment = LEFT if c == 3 else CENTER
    r += 1
ws_a.column_dimensions["A"].width = 11; ws_a.column_dimensions["B"].width = 24; ws_a.column_dimensions["C"].width = 95

# ───────────────────────── Resumen (portada) ─────────────────────────
ws_i = wb.create_sheet("Resumen", 0)
ws_i.sheet_view.showGridLines = False
ws_i.cell(row=1, column=1, value="CUADRANTE MENSUAL DE GARZONES — DIMANGO").font = F(bold=True, size=16, color="1F4E78")
ws_i.cell(row=2, column=1, value="Periodo: lunes 29 de junio a domingo 26 de julio de 2026 (4 semanas)").font = F(size=11)
lineas = [
    "", "HOJAS DE ESTE ARCHIVO:",
    "• Mall - Cuadrante     → grilla diaria de los 5 garzones base.",
    "• Playa - Cuadrante    → grilla diaria de los 6 base + 2 part-time.",
    "• Turnos               → leyenda de códigos, horarios y horas.",
    "• Resumen Horas        → horas por semana y total por trabajador (alerta si >42h).",
    "• Domingos Libres      → control de los domingos libres por trabajador.",
    "• Cobertura Diaria     → requerido vs asignado por día y local.",
    "• Alertas              → conflictos, ajustes y supuestos. LEER PRIMERO.",
    "", "REGLAS APLICADAS:",
    "• Base máx. 42h/semana. Modalidades 6x1 y 5x2.",
    "• 2 domingos libres/mes por trabajador (cumplido en Playa; parcial en Mall).",
    "• Mall: Lun-Jue 3 garzones · Vie-Dom 4 garzones.",
    "• Playa: 5 base/día (1 apertura, 1 intermedio, 3 cierre) + 1 apoyo PT Vie/Sáb/Dom.",
    "", "EQUIPOS:",
    "• Mall base: Deyanira, Khrisbell, Paula, Sofía, María.",
    "• Playa base: Jesús, Rubén, Marco, Nicolás, Fabián, Samuel.",
    "• Playa part-time: Jhony, Juan.",
    "", "⚠ Ver hoja 'Alertas': Mall logra ~1 domingo libre c/u; Playa con 6 base queda más ajustada (hasta 42h).",
]
r = 4
for ln in lineas:
    cell = ws_i.cell(row=r, column=1, value=ln)
    if ln.endswith(":"): cell.font = F(bold=True, size=11, color="1F4E78")
    elif ln.startswith("⚠"): cell.font = F(bold=True, size=11, color="9C0006")
    else: cell.font = F(size=10)
    r += 1
ws_i.column_dimensions["A"].width = 100

orden = ["Resumen", "Mall - Cuadrante", "Playa - Cuadrante", "Turnos",
         "Resumen Horas", "Domingos Libres", "Cobertura Diaria", "Alertas"]
wb._sheets.sort(key=lambda s: orden.index(s.title) if s.title in orden else 99)

OUT = "/Users/ricardovinet/whatsapp-agentkit/Cuadrante_Garzones_Dimango_Jun-Jul_2026.xlsx"
wb.save(OUT)
print("Guardado:", OUT)

# ───────────────────────── Validación ─────────────────────────
print("\n--- Validación ---")
ok = True
for w in PLAYA_FT:
    dl = sum(1 for di in DOM_IDX if playa_grid[w][di] == "LIB")
    if dl != 2: print(f"Playa {w}: domingos libres = {dl} <-- !"); ok = False
for w in MALL:
    dl = sum(1 for di in DOM_IDX if mall_grid[w][di] == "LIB")
    print(f"Mall {w}: domingos libres = {dl}")
for w in MALL:
    for wk in range(4):
        h = horas_semana(mall_grid[w], wk, MALL_H)
        if h != 42: print(f"Mall {w} sem{wk+1}: {h}h <-- ! (debe ser 42)"); ok = False
for i in range(DIAS):
    base = sum(1 for w in PLAYA_FT if playa_grid[w][i] != "LIB") + sum(1 for w in PLAYA_PT if pt_grid[w][i] == "PB")
    if base != 5: print(f"Playa {FECHAS[i].strftime('%d/%m')}: base={base} <-- !"); ok = False
    m = sum(1 for w in MALL if mall_grid[w][i] in ("AP", "IN", "CI"))
    if m != req_mall(WD[i]): print(f"Mall {FECHAS[i].strftime('%d/%m')}: cobertura={m} req={req_mall(WD[i])} <-- !"); ok = False
    ply = sum(1 for w in MALL if mall_grid[w][i] == "PLY")
    if ply: print(f"Mall {FECHAS[i].strftime('%d/%m')} ({DIAS_ABBR[WD[i]]}): {ply} a Playa (PLY)")
maxh = 0
for w in PLAYA_FT:
    for wk in range(4):
        maxh = max(maxh, horas_semana(playa_grid[w], wk, PLAYA_H))
print("Max horas/sem Playa FT:", maxh)
print("Cobertura y domingos:", "OK" if ok else "REVISAR")
